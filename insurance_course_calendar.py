
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import pyswarms as ps
from logger_config import logger



# Datos de ejemplo
# Datos de ejemplo

subjects = [
	'Intro a Seguros',         
	'Matemática Actuarial',    
	'Riesgos',                 
	'Legislación',             
	'Seguros de Vida',         
	'Seguros Generales',       
	'Gestión Comercial',       
	'Reaseguros'               
]
# prerequisites: (subject, prerequisite)
prerequisites = [
	('Matemática Actuarial', 'Intro a Seguros'),         
	('Riesgos', 'Matemática Actuarial'),                 
	('Legislación', 'Intro a Seguros'),                  
	('Seguros de Vida', 'Riesgos'),                      
	('Seguros Generales', 'Legislación'),                
	('Gestión Comercial', 'Seguros Generales'),          
	('Reaseguros', 'Seguros de Vida')                    
]
professors = ['Prof. A', 'Prof. B', 'Prof. C', 'Prof. D']
cohorts = ['C1', 'C2', 'C3', 'C4', 'C5']
# Asignar profesor fijo a cada materia
professor_by_subject = {
	'Intro a Seguros': 'Prof. A',
	'Matemática Actuarial': 'Prof. B',
	'Riesgos': 'Prof. C',
	'Legislación': 'Prof. D',
	'Seguros de Vida': 'Prof. A',
	'Seguros Generales': 'Prof. B',
	'Gestión Comercial': 'Prof. C',
	'Reaseguros': 'Prof. D'
}





logger.info('Starting configuration setup')
print_image_result = False  # Por defecto no imprime imagen
print_excel_result = True    # Por defecto sí imprime Excel
num_subjects = len(subjects)
num_professors = len(professors)
num_cohorts = len(cohorts)
num_years = 2
semesters_per_year = 2
start_year = 2026
num_semesters = num_years * semesters_per_year
logger.info(f'Config: subjects={num_subjects}, professors={num_professors}, cohorts={num_cohorts}, years={num_years}, semesters={num_semesters}')
max_classes_per_week = 2
max_subjects_per_day_professor = 2
max_subjects_per_day_cohort = 2 # cantidad máxima de días que puede asistir una cohorte por cuatrimestre (por defecto 1)
week_days = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
shifts = ['Mañana', 'Tarde']
# Bloques bloqueados: (día, turno)
blocked_slots = [('Viernes', 'Tarde'), ('Lunes', 'Mañana')]



# Cada materia debe asignarse a cuatrimestre, cohorte, día y turno (profesor fijo por materia)
# Vector de decisión: [materia1_cuatri, materia1_cohorte, materia1_dia, materia1_turno, materia2_cuatri, ...]
def decode_solution(x):
	logger.debug('Decoding solution vector')
	sol = []
	for i in range(num_subjects):
		for cohort_idx in range(num_cohorts):
			base = i*num_cohorts*4 + cohort_idx*4
			semester = int(round(x[base])) % num_semesters
			day = int(round(x[base+1])) % len(week_days)
			shift = int(round(x[base+2])) % len(shifts)
			# fixed professor by subject
			prof = professor_by_subject[subjects[i]]
			cohort = cohorts[cohort_idx]
			sol.append((subjects[i], semester, prof, cohort, week_days[day], shifts[shift]))
	logger.debug(f'Decoded solution length: {len(sol)}')
	return sol





# PSO con pyswarms
dim = num_subjects * num_cohorts * 4
max_iters = 200
bounds = (np.zeros(dim), np.ones(dim) * max(num_semesters, len(week_days), len(shifts)))


# Definición de la función objetivo antes de objective_pyswarms
def objective(x):
	logger.debug('Evaluating objective function')
	sol = decode_solution(x)
	penalty = 0
	# Restricción exacta: cada cohorte debe tener exactamente num_subjects / num_semesters materias por cuatrimestre
	if num_subjects % num_semesters == 0:
		expected_subjects = num_subjects // num_semesters
	else:
		expected_subjects = int(round(num_subjects / num_semesters))

	for cohort in cohorts:
		subjects_per_semester = [0]*num_semesters
		for _, semester, _, coh, _, _ in sol:
			if coh == cohort:
				subjects_per_semester[semester] += 1
		for semester_idx, count in enumerate(subjects_per_semester):
			if count != expected_subjects:
				penalty += 1e6 * abs(count - expected_subjects)
				#eturn float('inf')  # O raise Exception("Restricción violada")
	
	# Penalizar correlativas mal asignadas (por cohorte)
	for cohort in cohorts:
		semester_dict = {m: s for m, s, _, coh, _, _ in sol if coh == cohort}
		for subj, prereq in prerequisites:
			if semester_dict[subj] <= semester_dict[prereq]:
				penalty += 1000


	# Penalizar si no hay exactamente 2 materias por cuatrimestre por cohorte
	'''classes_per_cohort = {c: [0]*num_semesters for c in cohorts}
	for _, semester, _, cohort, _, _ in sol:
		classes_per_cohort[cohort][semester] += 1
	for cohort in cohorts:
		for semester in range(num_semesters):
			if classes_per_cohort[cohort][semester] != 2:
				penalty += 10000 * abs(classes_per_cohort[cohort][semester] - 2)'''

	# Penalizar bloques bloqueados (muy fuerte)
	for _, _, _, _, day, shift in sol:
		if (day, shift) in blocked_slots:
			penalty += 5000

	# Penalizar exceso de materias por día por cohorte y cuatrimestre
	for cohort in cohorts:
		for semester in range(num_semesters):
			subjects_per_day = {d: 0 for d in week_days}
			for _, s, _, coh, day, _ in sol:
				if coh == cohort and s == semester:
					subjects_per_day[day] += 1

			# Penalizar si hay más materias por día que lo permitido (profesor)
			for d in week_days:
				for prof in professors:
					subjects_prof_day = 0
					for _, s, p, coh, day, _ in sol:
						if coh == cohort and s == semester and day == d and p == prof:
							subjects_prof_day += 1
					if subjects_prof_day > max_subjects_per_day_professor:
						penalty += 50 * (subjects_prof_day - max_subjects_per_day_professor)

				
			# Penalizar si la cohorte asiste más días de los permitidos en el cuatrimestre
			attended_days = sum(1 for v in subjects_per_day.values() if v > 0)
			if attended_days > max_subjects_per_day_cohort:
				penalty += 200 * (attended_days - max_subjects_per_day_cohort)

	# Penalizar exceso de materias por día por profesor
	for prof in professors:
		for d in week_days:
			subjects_prof_day = 0
			for _, _, p, _, day, _ in sol:
				if p == prof and day == d:
					subjects_prof_day += 1
			if subjects_prof_day > max_subjects_per_day_professor:
				penalty += 100 * (subjects_prof_day - max_subjects_per_day_professor)

	# Penalizar exceso de materias por día por cohorte (en todo el calendario)
	for cohort in cohorts:
		for d in week_days:
			subjects_cohort_day = 0
			for _, _, _, coh, day, _ in sol:
				if coh == cohort and day == d:
					subjects_cohort_day += 1
			if subjects_cohort_day > max_subjects_per_day_cohort:
				penalty += 200 * (subjects_cohort_day - max_subjects_per_day_cohort)
	
	# Penalizar materias en el mismo bloque para la misma cohorte y cuatrimestre
	slots = set()
	for subj, semester, prof, cohort, day, shift in sol:
		key = (cohort, semester, day, shift)
		if key in slots:
			penalty += 100
		else:
			slots.add(key)

	# Penalizar desbalance de días
	days_count = {d: 0 for d in week_days}
	for _, _, _, _, day, _ in sol:
		days_count[day] += 1
	penalty += np.std(list(days_count.values())) * 	100

	# Penalizar si no se usan ambos turnos cada día
	for d in week_days:
		shifts_used = set()
		for _, _, _, _, day, shift in sol:
			if day == d:
				shifts_used.add(shift)
		if len(shifts_used) < len(shifts):
			# Penalizar si falta algún turno ese día
			penalty += 200 * (len(shifts) - len(shifts_used))
	return penalty

def objective_pyswarms(x):
	return np.array([objective(xi) for xi in x])

logger.info('Initializing PSO optimizer')
optimizer = ps.single.GlobalBestPSO(n_particles=60, dimensions=dim, options={'c1': 1.5, 'c2': 1.5, 'w': 0.7}, bounds=bounds)
logger.info('Starting PSO optimization')
cost, pos = optimizer.optimize(objective_pyswarms, iters=max_iters)
logger.info(f'Optimization finished. Best cost: {cost}')
convergencia = optimizer.cost_history
best_x = pos

# Mostrar convergencia
if print_image_result:
	logger.info('Plotting and saving convergence image')
	fig_conv, ax_conv = plt.subplots(figsize=(8,4))
	ax_conv.plot(convergencia)
	ax_conv.set_title('Convergencia PSO (pyswarms)')
	ax_conv.set_xlabel('Iteración')
	ax_conv.set_ylabel('Penalidad')
	ax_conv.grid()
	plt.tight_layout()
	plt.savefig('output_imgs/convergencia_pso.png', dpi=200)
	plt.show()



# Mostrar calendario por cohorte y por año (visual)
import matplotlib.dates as mdates
import datetime
sol = decode_solution(best_x)
colors = ['tab:blue','tab:orange','tab:green','tab:red']





# Consolidado: matriz día vs turno/cuatrimestre con todas las cohortes


# Mostrar cada cuatrimestre en una figura separada tipo calendario
if print_image_result:
	for semester_idx in range(num_semesters):
		table_data = []
		for shift in shifts:
			row = []
			for day in week_days:
				cell = []
				for m, s, prof, coh, d, t in sol:
					if s == semester_idx and d == day and t == shift:
						cell.append(f'{m} - {coh} ({prof})')
				if cell:
					cell_str = ('\n').join(cell)
				else:
					cell_str = ''
				row.append(cell_str)
			table_data.append(row)
		row_labels = [f'{t}' for t in shifts]
		df = pd.DataFrame(table_data, columns=week_days, index=row_labels)
		fig, ax = plt.subplots(figsize=(14,5))
		ax.axis('off')
		tbl = ax.table(cellText=df.values, rowLabels=df.index, colLabels=df.columns, loc='center', cellLoc='left', edges='open')
		tbl.auto_set_font_size(False)
		tbl.set_fontsize(16)
		for (row, col), cell in tbl.get_celld().items():
			cell.set_width(0.22)
			cell.set_height(0.18)
			cell.set_linewidth(2)
			if row == 0 or col == -1:
				cell.set_facecolor('#dbeafe')
				cell.set_text_props(va='top', ha='left', fontsize=16)
			elif (row+col)%2 == 0:
				cell.set_facecolor('#f1f5f9')
				cell.set_text_props(va='top', ha='left', fontsize=11)
			else:
				cell.set_facecolor('#e0e7ef')
				cell.set_text_props(va='top', ha='left', fontsize=11)
		plt.title(f'Calendario Consolidado - Semestre {semester_idx+1}', fontsize=22, fontweight='bold')
		plt.tight_layout()
		plt.savefig(f'output_imgs/calendario_consolidado_semestre{semester_idx+1}.png', dpi=200)
		plt.show()





if print_image_result:
	for year_idx in range(num_years):
		year = start_year + year_idx
		for cohort in cohorts:
			for local_semester in range(semesters_per_year):
				table_data = []
				for shift in shifts:
					row = []
					for day in week_days:
						cell = []
						for subj, semester, prof, coh, d, t in sol:
							if coh == cohort and semester // semesters_per_year == year_idx and semester % semesters_per_year == local_semester and d == day and t == shift:
								cell.append(f'{subj} ({prof})')
						if cell:
							cell_str = ('\n').join(cell)
						else:
							cell_str = ''
						row.append(cell_str)
					table_data.append(row)
				row_labels = [f'{t}' for t in shifts]
				df = pd.DataFrame(table_data, columns=week_days, index=row_labels)
				fig, ax = plt.subplots(figsize=(12,5))
				ax.axis('off')
				tbl = ax.table(cellText=df.values, rowLabels=df.index, colLabels=df.columns, loc='center', cellLoc='left', edges='open')
				tbl.auto_set_font_size(False)
				tbl.set_fontsize(16)
				for (row, col), cell in tbl.get_celld().items():
					cell.set_width(0.22)
					cell.set_height(0.18)
					cell.set_linewidth(2)
					if row == 0 or col == -1:
						cell.set_facecolor('#dbeafe')
						cell.set_text_props(va='top', ha='left', fontsize=16)
					elif (row+col)%2 == 0:
						cell.set_facecolor('#f1f5f9')
						cell.set_text_props(va='top', ha='left', fontsize=11)
					else:
						cell.set_facecolor('#e0e7ef')
						cell.set_text_props(va='top', ha='left', fontsize=11)
				plt.title(f'Calendar {cohort} - Year {year} - Semester {local_semester+1}', fontsize=20, fontweight='bold')
				plt.tight_layout()
				plt.savefig(f'output_imgs/calendar_{cohort}_year{year}_semester{local_semester+1}.png', dpi=200)
				plt.show()





if print_excel_result:
	logger.info('Exporting results to Excel')
	from openpyxl import Workbook
	from openpyxl.utils import get_column_letter
	wb = Workbook()
	ws_consolidated = wb.create_sheet(title='Consolidated')
	ws_consolidated.append(['Year', 'Semester/Shift'] + week_days)
	max_col_len = [0]*(len(week_days)+2)
	for semester in range(num_semesters):
		for shift in shifts:
			year = start_year + (semester // semesters_per_year)
			row = [str(year), f'Sem.{(semester%semesters_per_year)+1} {shift}']
			for idx, day in enumerate(week_days):
				cell = []
				for subj, s, prof, coh, d, t in sol:
					if s == semester and d == day and t == shift:
						cell.append(f'{subj} :: {coh} ({prof})')
				if cell:
					cell_str = ('\n').join(cell)
				else:
					cell_str = ''
				row.append(cell_str)
				max_col_len[idx+2] = max(max_col_len[idx+2], len(cell_str))
			max_col_len[0] = max(max_col_len[0], len(row[0]))
			max_col_len[1] = max(max_col_len[1], len(row[1]))
			ws_consolidated.append(row)
	# Auto-adjust column width and row height
	for i in range(1, len(week_days)+3):
		ws_consolidated.column_dimensions[get_column_letter(i)].width = max(22, min(60, max_col_len[i-1]*1.2))
	for i in range(1, ws_consolidated.max_row+1):
		ws_consolidated.row_dimensions[i].height = 40

	# One sheet per cohort, all semesters and years together
	for cohort in cohorts:
		ws = wb.create_sheet(title=f'{cohort}')
		ws.append(['Year', 'Semester/Shift'] + week_days)
		max_col_len = [0]*(len(week_days)+2)
		for semester in range(num_semesters):
			year = start_year + (semester // semesters_per_year)
			for shift in shifts:
				row = [str(year), f'Sem.{(semester%semesters_per_year)+1} {shift}']
				for idx, day in enumerate(week_days):
					cell = []
					for subj, s, prof, coh, d, t in sol:
						if coh == cohort and s == semester and d == day and t == shift:
							cell.append(f'{subj} ({prof})')
					if cell:
						cell_str = ('\n').join(cell)
					else:
						cell_str = ''
					row.append(cell_str)
					max_col_len[idx+2] = max(max_col_len[idx+2], len(cell_str))
				max_col_len[0] = max(max_col_len[0], len(row[0]))
				max_col_len[1] = max(max_col_len[1], len(row[1]))
				ws.append(row)
		# Auto-adjust column width and row height
		for i in range(1, len(week_days)+3):
			ws.column_dimensions[get_column_letter(i)].width = max(22, min(60, max_col_len[i-1]*1.2))
		for i in range(1, ws.max_row+1):
			# Adjust row height according to number of lines
			max_lines = 1
			for cell in ws[i]:
				val = str(cell.value) if cell.value else ''
				max_lines = max(max_lines, val.count('\n')+1)
			ws.row_dimensions[i].height = 20 + 15*max_lines
	wb.remove(wb['Sheet'])
	wb.save('output_excel/calendar_subjects.xlsx')
	logger.info('Excel file saved to output_excel/calendar_subjects.xlsx')
